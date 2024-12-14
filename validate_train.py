import pickle
from sklearn.metrics import accuracy_score, classification_report
import pandas as pd
import time
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook

# โหลดข้อมูลใบหน้า
start_time = time.time()
# Load the test data and model
with open("models/train_test_data.pkl", "rb") as f:
    X_train, X_test, y_train, y_test = pickle.load(f)

with open("models/face_recognition_model.pkl", "rb") as f:
    clf = pickle.load(f)

# ทดสอบโมเดล
print("Testing the model...")
y_pred = clf.predict(X_test)
print("Accuracy:", accuracy_score(y_test, y_pred))
print(classification_report(y_test, y_pred))
print(f"Train classification report completed in {time.time() - start_time:.2f} seconds")

print("Plots saved and interpretations appended to classification report.")
# Generate classification report as a dictionary
report_dict = classification_report(y_test, y_pred, output_dict=True)

# Convert dictionary into the format provided
data = {
    "precision": [],
    "recall": [],
    "f1-score": [],
    "support": []
}

labels = list(report_dict.keys())[:-3]  # Skip 'accuracy', 'macro avg', 'weighted avg'

for label in labels:
    metrics = report_dict[label]
    data["precision"].append(metrics["precision"])
    data["recall"].append(metrics["recall"])
    data["f1-score"].append(metrics["f1-score"])
    data["support"].append(metrics["support"])

accuracy = report_dict["accuracy"]

sumSupport = sum(data["support"])
# Add Accuracy to the DataFrame
data["precision"].append(None)
data["recall"].append(None)
data["f1-score"].append(accuracy)
data["support"].append(sumSupport)

# Add the overall metrics ('macro avg' and 'weighted avg')
for avg in ["macro avg", "weighted avg"]:
    data["precision"].append(report_dict[avg]["precision"])
    data["recall"].append(report_dict[avg]["recall"])
    data["f1-score"].append(report_dict[avg]["f1-score"])
    data["support"].append(sumSupport)  # Overall metrics don't have support

# Convert to DataFrame for visualization
labels.extend(["accuracy", "macro avg", "weighted avg"])  # Add 'accuracy' to labels
df = pd.DataFrame(data, index=labels)
print(df)

# Plot 1: Precision by Class
plt.figure(figsize=(10, 6))
(df[:-3]["precision"] * 100).plot(kind="bar", color="skyblue", edgecolor="black")
plt.title("Precision by Class (in %)", fontsize=14)
plt.ylabel("Precision (%)", fontsize=12)
plt.xlabel("Class", fontsize=12)
plt.xticks(rotation=45, ha="right")
plt.ylim(90, 100)  # Percent range between 95 and 100
plt.tight_layout()
plt.savefig("./reports/train/1_precision_by_class_percent.png")

# Plot 2: Recall by Class
plt.figure(figsize=(10, 6))
(df[:-3]["recall"] * 100).plot(kind="bar", color="lightgreen", edgecolor="black")
plt.title("Recall by Class (in %)", fontsize=14)
plt.ylabel("Recall (%)", fontsize=12)
plt.xlabel("Class", fontsize=12)
plt.xticks(rotation=45, ha="right")
plt.ylim(90, 100)  # Percent range between 95 and 100
plt.tight_layout()
plt.savefig("./reports/train/2_recall_by_class_percent.png")

# Plot 3: F1-Score by Class
plt.figure(figsize=(10, 6))
(df[:-3]["f1-score"] * 100).plot(kind="bar", color="salmon", edgecolor="black")
plt.title("F1-Score by Class (in %)", fontsize=14)
plt.ylabel("F1-Score (%)", fontsize=12)
plt.xlabel("Class", fontsize=12)
plt.xticks(rotation=45, ha="right")
plt.ylim(90, 100)  # Percent range between 95 and 100
plt.tight_layout()
plt.savefig("./reports/train/3_f1_score_by_class_percent.png")

filtered_support_values = df[:-3]["support"].dropna()
filtered_class_labels = filtered_support_values.index
filtered_precision_values = df.loc[filtered_class_labels, "precision"].dropna()

# Generate labels with class name, support, and precision
pie_labels = [
    f"{cls} ({int(support)}) P: {precision * 100:.2f}%"
    for cls, support, precision in zip(
        filtered_class_labels, filtered_support_values, filtered_precision_values
    )
]

# Create pie chart with a clear template
plt.figure(figsize=(10, 8))
plt.pie(
    filtered_support_values,
    labels=pie_labels,
    autopct="%1.1f%%",
    startangle=90,
    colors=plt.cm.Paired.colors,
    radius=0.8,
    wedgeprops={"edgecolor": "black"},
    textprops={"fontsize": 12},
)
plt.title("Support and Precision Distribution by Class", fontsize=16, fontweight="bold")
plt.tight_layout()

# Save the pie chart
pie_chart_path = "./reports/train/4_support_precision_pie_chart.png"
plt.savefig(pie_chart_path)

# Extract macro avg and weighted avg values for precision, recall, and f1-score
averages = df.loc[["macro avg", "weighted avg"], ["precision", "recall", "f1-score"]]

# Convert to percentage for better readability
averages_percentage = averages * 100

# Plotting a grouped bar chart
fig, ax = plt.subplots(figsize=(10, 6))

# Create the bar chart
averages_percentage.plot(
    kind="bar",
    ax=ax,
    color=["skyblue", "lightgreen", "salmon"],
    edgecolor="black",
    width=0.8
)

# Add titles and labels
ax.set_title("Macro Average vs Weighted Average (in %)", fontsize=16, fontweight="bold")
ax.set_ylabel("Percentage (%)", fontsize=12)
ax.set_xlabel("Average Type", fontsize=12)
ax.set_xticks([0, 1])  # Positions for macro avg and weighted avg
ax.set_xticklabels(["Macro Avg", "Weighted Avg"], rotation=0, fontsize=12)
plt.ylim(90, 100)  # Percent range between 95 and 100
ax.legend(title="Metric", fontsize=10, title_fontsize=12)

# Add value labels above the bars
for container in ax.containers:
    ax.bar_label(container, fmt="%.2f%%", label_type="edge", fontsize=10, padding=3)

# Adjust layout for better readability
plt.tight_layout()

# Save the bar chart
bar_chart_path = "./reports/train/5_macro_weighted_avg_bar_chart.png"
plt.savefig(bar_chart_path)

plt.close()

print("Plots have been saved successfully.")

# Save the DataFrame to an Excel file
report_file_path = "reports/train/classification_report_train.xlsx"
df.to_excel(report_file_path, index=True)

# Load the existing Excel file
workbook = load_workbook(report_file_path)
sheet = workbook.active

# Add spacing (2 rows) below the current content
max_row = sheet.max_row
sheet.cell(row=max_row + 2, column=1).value = "Description"

interpretation = """
การตีความผลการประเมิน (Interpretation of Metrics):

1. Precision (ความแม่นยำเชิงบวก): เป็นการวัดสัดส่วนของการทำนายที่ถูกต้องในกลุ่มตัวอย่างที่ถูกทำนายว่าเป็นคลาสนั้น ๆ
   - Precision สูงหมายถึงมีการทำนายผิดพลาด (False Positive) น้อย
   - ตัวอย่างเช่น คลาส "Ball" มีค่า Precision ต่ำกว่าคลาสอื่น ๆ (0.956) เนื่องจากมีการทำนายผิดพลาดบางส่วน

2. Recall (การครอบคลุม): เป็นการวัดสัดส่วนของตัวอย่างที่ถูกทำนายถูกต้องเมื่อเทียบกับตัวอย่างทั้งหมดในคลาสนั้น
   - Recall สูงหมายถึงมีการทำนายตัวอย่างที่ควรจะเป็นคลาสนั้นพลาด (False Negative) น้อย
   - ตัวอย่างเช่น คลาส "Nik" มีค่า Recall ลดลงเล็กน้อย (0.979) เพราะมีบางตัวอย่างที่ไม่ได้ถูกทำนายอย่างถูกต้อง

3. F1-Score: เป็นค่าเฉลี่ยเชิงฮาร์มอนิกของ Precision และ Recall เพื่อหาค่าที่สมดุลระหว่างทั้งสอง
   - ค่า F1-Score ที่สมบูรณ์แบบ 1.0 สำหรับหลายคลาสบ่งบอกถึงประสิทธิภาพที่ยอดเยี่ยม
   - ตัวอย่างเช่น คลาส "Mile" มีค่า F1-Score อยู่ที่ 0.994 แสดงว่ามีข้อผิดพลาดน้อยมาก

4. Accuracy (ความถูกต้อง): แสดงถึงสัดส่วนของตัวอย่างทั้งหมดที่ถูกทำนายอย่างถูกต้อง
   - ค่า Accuracy สำหรับโมเดลนี้คือ 99.63% แสดงถึงประสิทธิภาพโดยรวมที่ยอดเยี่ยม

ค่าเฉลี่ยแบบ Macro และ Weighted:
- Macro avg: คำนวณค่าเฉลี่ยของ Precision, Recall และ F1-Score แบบไม่มีการถ่วงน้ำหนัก ซึ่งทุกคลาสจะมีน้ำหนักเท่ากัน
- Weighted avg: คำนวณค่าเฉลี่ยโดยคำนึงถึงจำนวนตัวอย่างในแต่ละคลาส (support) ซึ่งให้ผลลัพธ์ที่สะท้อนประสิทธิภาพโดยรวมได้ดีกว่า
"""

# Write interpretation line by line below the table
for i, line in enumerate(interpretation.strip().split("\n")):
    sheet.cell(row=max_row + 3 + i, column=1).value = line

# Save the updated Excel file
workbook.save(report_file_path)

print(f"Saved as: {report_file_path}")



